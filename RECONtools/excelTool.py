#!/usr/bin/env python
import openpyxl
import sys, re
import Series
# Gathering excel info
def importColumns(xl_template):
    '''Builds/returns a list of columns found in template'''
    columns = []
    for col in range( xl_template.ncols ):
        columns.append( str(xl_template.cell(0,col).value) )
    return columns

def processExp(exp): #===
    # Need to cap at end? $
    '''Converts filter from RECONSTRUCT format to python regexp format'''
    exp = exp.replace('*', '.')
    exp = exp.replace('?', '[a-z]') #=== "? matches any single character" (numbers included?)
    exp = exp.replace('#', '[0-9]')
    return re.compile( str(exp),re.I ) # re.I (Ignore Case)

def buildFilterBank(list_of_expressions):
    '''Turns a list of expression into a list of regular expressions for filtering'''
    bank = []
    for exp in list_of_expressions:
        exp = processExp( exp )
        bank.append( exp )
    return bank

def buildProtList(series, filterBank):
    '''Returns a list of contour names from series, specified by filterBank'''
    protList = []
    for section in series.sections:
        for contour in [cont.name for cont in section.contours]:
            for exp in filterBank:
                if exp.match(contour) != None:
                    protList.append(contour)
    return list(set(protList))

def buildProtDict(protList):
    '''Creates a dictionary of protrusion names, each associated with a list of objects representing data'''
    protDict = {}
    for prot in protList:
        protDict[prot] = []
    return protDict


# Creating excel docs
def writeColumns(worksheet, list_of_columns):
    '''Writes columns to specified worksheet.'''
    for col in range( len(list_of_columns)):
        worksheet.cell(row=0, column=col).value = str( list_of_columns[col] )
        
def addProtsToSheet(worksheet, protrusion_list): #===
    '''Adds protrusions in protrusion_list to the appropriate column in worksheet.'''
    return

def addDendriteSheet(workbook, series_name, xl_template, dendrite):
    '''Adds a worksheet to workbook for the specific dendrite. Formatted to template.'''
    worksheet = workbook.create_sheet( title=series_name+' '+dendrite )
    writeColumns(worksheet, importColumns(xl_template))

def addDendriteSheets(workbook, series_name, xl_template, dendrite_list):
    '''Adds a sheet to workbook for each dendrite of dendrite_list.'''
    for dendrite in dendrite_list:
        addDendriteSheet(workbook, series_name, xl_template, dendrite)
    